
# coding: utf-8

# In[130]:

#import openpyxl 
from openpyxl import *
from tkinter import *
from tkinter import messagebox
# from tkinter.ttk import *
import csv
import os
import openpyxl


# Creating xlsx file for database
with open('py_project_db.csv', 'wb') as csvfile:
    filewriter = csv.writer(csvfile, delimiter=',',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)

wb = openpyxl.Workbook()
ws = wb.active

with open('py_project_db.csv') as f:
    reader = csv.reader(f, delimiter=':')
    for row in reader:
        ws.append(row)

wb.save('py_project_db.xlsx')

os.remove('py_project_db.csv')

# opening the excel file 
wb = load_workbook('py_project_db.xlsx') 

# create the sheet object 
sheet = wb.active 


def py_project_db(): 
    
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 8
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 8
    sheet.column_dimensions['I'].width = 8
    sheet.column_dimensions['K'].width = 8
    sheet.column_dimensions['L'].width = 8
    sheet.column_dimensions['M'].width = 8
    sheet.column_dimensions['N'].width = 8
    sheet.column_dimensions['O'].width = 8
    sheet.column_dimensions['P'].width = 8
    sheet.column_dimensions['Q'].width = 8
    sheet.column_dimensions['R'].width = 8
    sheet.column_dimensions['S'].width = 8
    sheet.column_dimensions['T'].width = 8
    sheet.column_dimensions['U'].width = 8
    
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Registration no."
    sheet.cell(row=1, column=3).value = "Gender"
    sheet.cell(row=1, column=4).value = "Branch"
    sheet.cell(row=1, column=5).value = "Semester"
    sheet.cell(row=1, column=6).value = "Contact Number"
    sheet.cell(row=1, column=7).value = "Email id"
    sheet.cell(row=1, column=8).value = "Hostel"
    sheet.cell(row=1, column=9).value = "Art"
    sheet.cell(row=1, column=10).value = "NCC"
    sheet.cell(row=1, column=11).value = "Sports"
    sheet.cell(row=1, column=12).value = "Tech"
    sheet.cell(row=1, column=13).value = "Skills"
    sheet.cell(row=1, column=14).value = "Event Mngmt"
    sheet.cell(row=1, column=15).value = "LLC"
    sheet.cell(row=1, column=16).value = "METAL"
    sheet.cell(row=1, column=17).value = "SPADE"
    sheet.cell(row=1, column=18).value = "ECHO"
    sheet.cell(row=1, column=19).value = "TRANQUIL"
    sheet.cell(row=1, column=20).value = "CLOUDBUGS"

    
    
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the course_field box 
    regd_field.focus_set() 


# # Function to set focus 
# def focus2(event): 
#     # set focus on the sem_field box 
#     gender_field.focus_set() 
    

# Function to set focus 
def focus3(event): 
    # set focus on the form_no_field box 
    branch_field.focus_set() 


# Function to set focus 
def focus4(event): 
    # set focus on the contact_no_field box 
        
    semester_field.focus_set() 


# Function to set focus 
def focus5(event): 
    # set focus on the email_id_field box 
    contact_field.focus_set() 
    

# Function to set focus 
def focus6(event): 
    # set focus on the address_field box 
    email_field.focus_set() 

# Function to set focus 
def focus7(event): 
    # set focus on the address_field box 
    hostel_field.focus_set()

# # Function to set focus 
# def focus8(event): 
# # set focus on the address_field box 
#     interest_field.focus_set() 
    
# # Function to set focus 
# def focus9(event): 
#     # set focus on the address_field box 
#     club_field.focus_set() 
    
def focus10(event): 
    # set focus on the address_field box 
    submit.focus_set()     
        
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
    
    # clear the content of text entry box 
    name_field.delete(0, END) 
    regd_field.delete(0, END) 
#     gender_field.delete(0, END) 
#     branch_field.delete(0, END) 
    semester_field.delete(0, END) 
    contact_field.delete(0, END) 
    email_field.delete(0, END) 
    hostel_field.delete(0, END)    

# Function to take data from GUI 
# window and write to an excel file 
def insert(): 

    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        regd_field.get() == "" and
        semester_field.get() == "" and
        gender_field.get() == "" and
        contact_field.get() == "" and
        email_field.get() == "" and
        branch_field.get() == ""): 
        print("empty input") 

    else: 

        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 

        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = regd_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = gender_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = branch_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = semester_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = contact_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = email_field.get() 
        sheet.cell(row=current_row + 1, column=8).value = hostel_field.get() 
        sheet.cell(row=current_row + 1, column=9).value = chk_state.get() 
        sheet.cell(row=current_row + 1, column=10).value = chk2_state.get() 
        sheet.cell(row=current_row + 1, column=11).value = chk3_state.get() 
        sheet.cell(row=current_row + 1, column=12).value = chk4_state.get() 
        sheet.cell(row=current_row + 1, column=13).value = chk5_state.get() 
        sheet.cell(row=current_row + 1, column=14).value = chk6_state.get() 
        sheet.cell(row=current_row + 1, column=15).value = chk7_state.get()         
        sheet.cell(row=current_row + 1, column=16).value = chk8_state.get()         
        sheet.cell(row=current_row + 1, column=17).value = chk9_state.get()         
        sheet.cell(row=current_row + 1, column=18).value = chk10_state.get()         
        sheet.cell(row=current_row + 1, column=19).value = chk11_state.get() 
        sheet.cell(row=current_row + 1, column=20).value = chk12_state.get()         

        # save the file 
        wb.save('py_project_db.xlsx') 

        # set focus on the name_field box 
        name_field.focus_set() 

        # call the clear() function 
        clear() 


# Driver code 
if __name__ == "__main__": 

    # create a GUI window 
    root = Tk() 

    # set the background colour of GUI window 
    root.configure(background='aqua') 

    # set the title of GUI window 
    root.title("Registration form") 
    
    # set the configuration of GUI window 
    root.geometry("1180x850") 

    py_project_db() 

    # create a Form label 
    heading = Label(root, text="WELCOME", font=("Algerian", 50),background="Pink", bg="light green") 

    # create a Name label 
    name = Label(root, text="Name         :", bg="light yellow", font=('Berlin Sans FB', 18)) 

    # create a Course label 
    regd = Label(root, text="Regd. No.  :", bg="yellow", font=('Berlin Sans FB', 18)) 

    # create a Semester label 
    gender = Label(root, text="Gender      :", bg="light yellow", font=('Berlin Sans FB', 18)) 

    # create a Form No. lable 
    branch = Label(root, text="Branch       :", bg="yellow", font=('Berlin Sans FB', 18)) 

    # create a Contact No. label 
    semester = Label(root, text="Semester   :", bg="light yellow", font=('Berlin Sans FB', 18)) 

    # create a Email id label 
    contact = Label(root, text="Contact No.:", bg="yellow", font=('Berlin Sans FB', 18)) 

    # create a address label 
    email = Label(root, text="E-mail        :", bg="light yellow", font=('Berlin Sans FB', 18)) 
    
    # create a Email id label 
    hostel = Label(root, text="Hostel(if Y) :", bg="yellow", font=('Berlin Sans FB', 18)) 

    # create a address label 
    interest = Label(root, text="Interested in -", bg="light yellow", font=('Berlin Sans FB', 18)) 
    
    # create a Email id label 
    club = Label(root, text="Preferences -", bg="yellow", font=('Berlin Sans FB', 18)) 

    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=3) 
    name.grid(row=1, column=0) 
    regd.grid(row=2, column=0) 
    gender.grid(row=3, column=0) 
    branch.grid(row=4, column=0)
    semester.grid(row=5, column=0) 
    contact.grid(row=6, column=0) 
    email.grid(row=7, column=0) 
    hostel.grid(row=8, column=0) 
    interest.grid(row=9, column=0) 
    club.grid(row=11, column=0)

    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    regd_field = Entry(root) 
    semester_field = Entry(root)
    contact_field = Entry(root) 
    email_field = Entry(root) 
    branch_field = ttk.Combobox(root,values= ("Engineering", "Fashion designing", "Management", "Agriculture", "Forensics", "Physics", "Hospitality", "Architecture", "Airlines"))
    branch_field.current(0)
    
    hostel_field = Entry(root)
    gender_field = StringVar()
    male = Radiobutton(root,text='Male', font=('Berlin Sans FB', 11), value='M', var=gender_field)
    female = Radiobutton(root,text='Female', font=('Berlin Sans FB', 11), value='F', var=gender_field)
    other_g = Radiobutton(root,text='Other', font=('Berlin Sans FB', 11), value='O', var=gender_field)
 
    chk_state = IntVar()
    chk_state.set(0) #set check state
    chk = Checkbutton(root, text='Art/Creativity', font=('Berlin Sans FB', 12), var=chk_state, onvalue=1, offvalue=0) 
    chk2_state = IntVar()
    chk2_state.set(0) #set check state
    chk2 = Checkbutton(root, text='NCC                      `', font=('Berlin Sans FB', 12), var=chk2_state, onvalue=1, offvalue=0)
    chk3_state = IntVar()
    chk3_state.set(0) #set check state
    chk3 = Checkbutton(root, text='Sports                     `', font=('Berlin Sans FB', 12), var=chk3_state, onvalue=1, offvalue=0)
    chk4_state = IntVar()
    chk4_state.set(0) #set check state
    chk4 = Checkbutton(root, text='Technology `', font=('Berlin Sans FB', 12), var=chk4_state, onvalue=1, offvalue=0) 
    chk5_state = IntVar() 
    chk5_state.set(0) #set check state
    chk5 = Checkbutton(root, text='Interpersonal Skills', font=('Berlin Sans FB', 12), var=chk5_state, onvalue=1, offvalue=0)
    chk6_state = IntVar()
    chk6_state.set(0) #set check state
    chk6 = Checkbutton(root, text='Event Management', font=('Berlin Sans FB', 12), var=chk6_state, onvalue=1, offvalue=0)
    
    
    chk7_state = IntVar()
    chk7_state.set(0) #set check state
    chk7 = Checkbutton(root, text='LEADER\'S LANTERN', font=('Berlin Sans FB', 12), var=chk7_state, onvalue=1, offvalue=0) 
    chk8_state = IntVar() 
    chk8_state.set(0) #set check state
    chk8 = Checkbutton(root, text='METAL            `', font=('Berlin Sans FB', 12), var=chk8_state, onvalue=1, offvalue=0)
    chk9_state = IntVar()
    chk9_state.set(0) #set check state
    chk9 = Checkbutton(root, text='SPADE                       `', font=('Berlin Sans FB', 12), var=chk9_state, onvalue=1, offvalue=0)
    chk10_state = IntVar()
    chk10_state.set(0) #set check state
    chk10 = Checkbutton(root, text='ECHO              `', font=('Berlin Sans FB', 12), var=chk10_state, onvalue=1, offvalue=0) 
    chk11_state = IntVar() 
    chk11_state.set(0) #set check state
    chk11 = Checkbutton(root, text='TRANQUIL                  `', font=('Berlin Sans FB', 12), var=chk11_state, onvalue=1, offvalue=0)
    chk12_state = IntVar()  
    chk12_state.set(0) #set check state
    chk12 = Checkbutton(root, text='CLOUD BUGS', font=('Berlin Sans FB', 12), var=chk12_state, onvalue=1, offvalue=0)
    

    # bind method of widget is used for 
    # the binding the function with the events 

    # whenever the enter key is pressed 
    # then call the focus1 function 
    name_field.bind("<Return>", focus1) 

    # whenever the enter key is pressed 
    # then call the focus2 function 
    regd_field.bind("<Return>", focus2) 

    # whenever the enter key is pressed 
    # then call the focus4 function 
    branch_field.bind("<Return>", focus4) 

    # whenever the enter key is pressed 
    # then call the focus5 function 
    semester_field.bind("<Return>", focus5) 

     # whenever the enter key is pressed 
    # ten call the focus6 function 
    contact_field.bind("<Return>", focus6) 

    # whenever the enter key is pressed 
    # then call the focus7 function 
    email_field.bind("<Return>", focus7) 
    
    # whenever the enter key is pressed 
    # then call the focus8 function 
    hostel_field.bind("<Return>", focus8) 

        
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    name_field.grid(row=1, column=2, ipadx="80", ipady="5") 
    regd_field.grid(row=2, column=2, ipadx="80", ipady="5") 
    male.grid(column=2, row=3)
    female.grid(column=3, row=3)
    other_g.grid(column=4, row=3)
    branch_field.grid(row=4, column=2, ipadx="70") 
    semester_field.grid(row=5, column=2, ipadx="80", ipady="5") 
    contact_field.grid(row=6, column=2, ipadx="80", ipady="5") 
    email_field.grid(row=7, column=2, ipadx="80", ipady="5") 
    hostel_field.grid(row=8, column=2, ipadx="80", ipady="5") 
    
    chk.grid(column=2, row=9)
    chk2.grid(column=3, row=9)
    chk3.grid(column=5, row=9)                                                              
    chk4.grid(column=2, row=10)
    chk5.grid(column=3, row=10)
    chk6.grid(column=5, row=10)
    
    chk7.grid(column=2, row=11)
    chk8.grid(column=3, row=11)
    chk9.grid(column=2, row=12)                                                              
    chk10.grid(column=3, row=12)
    chk11.grid(column=2, row=13)
    chk12.grid(column=3, row=13)
    
    # call excel function 
    py_project_db() 

    # create a Submit Button and place into the root window 

    submit = Button(root, text="Submit", fg="Black", bg="Red", font=('Berlin Sans FB', 25), command=insert) 
    submit.grid(row=25, column=3) 

  
    # start the GUI 
    root.mainloop() 

